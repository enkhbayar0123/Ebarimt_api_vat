"""
eBarimt VAT Checker - Web App
==============================
Run:   python app.py
Then open your browser to:  http://localhost:5000
"""

import sys
import os
import json
import time
import threading
import webbrowser
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs, unquote, quote
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError

# ── Auto-install openpyxl if needed ──────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

from io import BytesIO
from datetime import datetime

PORT = int(os.environ.get("PORT", 5000))
API_URL = "https://api.ebarimt.mn/api/info/check/getTinInfo?regNo={}"

# ── In-memory job state ───────────────────────────────────────────────────────
job = {
    "running": False,
    "results": [],   # list of {reg, tin, status}
    "total": 0,
    "done": 0,
}
job_lock = threading.Lock()

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>eBarimt VAT Checker</title>
<style>
:root {
  --blue:#1a56db; --blue-dark:#1342b0;
  --green-bg:#f0fdf4; --green:#15803d; --green-bd:#bbf7d0;
  --red-bg:#fef2f2;   --red:#b91c1c;   --red-bd:#fecaca;
  --yellow-bg:#fefce8;--yellow:#a16207;--yellow-bd:#fde68a;
  --gray-50:#f9fafb; --gray-100:#f3f4f6; --gray-200:#e5e7eb;
  --gray-400:#9ca3af; --gray-600:#4b5563; --gray-800:#1f2937;
  --radius:8px; --radius-lg:12px;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#f1f5f9;color:var(--gray-800);min-height:100vh}

/* ── Top bar ── */
.topbar{background:#fff;border-bottom:1px solid var(--gray-200);padding:0 32px;height:56px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:10}
.topbar-brand{display:flex;align-items:center;gap:10px;font-size:15px;font-weight:600;color:var(--gray-800)}
.topbar-brand svg{color:var(--blue)}
.topbar-sub{font-size:12px;color:var(--gray-400);font-weight:400}

/* ── Layout ── */
.page{max-width:960px;margin:0 auto;padding:28px 24px}
.grid{display:grid;grid-template-columns:320px 1fr;gap:20px;align-items:start}
@media(max-width:720px){.grid{grid-template-columns:1fr}}

/* ── Cards ── */
.card{background:#fff;border:1px solid var(--gray-200);border-radius:var(--radius-lg);padding:20px}
.card-title{font-size:13px;font-weight:600;color:var(--gray-600);text-transform:uppercase;letter-spacing:.05em;margin-bottom:14px}

/* ── Input panel ── */
.field-label{font-size:13px;color:var(--gray-600);margin-bottom:6px;display:block}
textarea{width:100%;height:180px;resize:vertical;font-family:monospace;font-size:12px;border:1px solid var(--gray-200);border-radius:var(--radius);padding:10px;color:var(--gray-800);background:#fff;outline:none;line-height:1.6}
textarea:focus{border-color:var(--blue);box-shadow:0 0 0 3px rgba(26,86,219,.1)}
.or-divider{text-align:center;font-size:12px;color:var(--gray-400);margin:10px 0}
.file-zone{border:1.5px dashed var(--gray-200);border-radius:var(--radius);padding:16px;text-align:center;cursor:pointer;transition:border-color .15s}
.file-zone:hover{border-color:var(--blue)}
.file-zone input{display:none}
.file-zone-text{font-size:13px;color:var(--gray-400)}
.file-zone-text b{color:var(--blue)}
#fileNameTag{font-size:12px;color:var(--green);margin-top:4px}
#countTag{font-size:12px;color:var(--gray-400);margin-top:8px}

.settings-row{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:14px}
.settings-row label{font-size:12px;color:var(--gray-600);display:flex;flex-direction:column;gap:4px}
select{padding:7px 8px;border:1px solid var(--gray-200);border-radius:var(--radius);font-size:13px;color:var(--gray-800);background:#fff;cursor:pointer;outline:none}
select:focus{border-color:var(--blue)}

.btn{width:100%;margin-top:14px;padding:10px;border-radius:var(--radius);border:none;font-size:14px;font-weight:600;cursor:pointer;display:flex;align-items:center;justify-content:center;gap:8px;transition:background .15s}
.btn-start{background:var(--blue);color:#fff}
.btn-start:hover{background:var(--blue-dark)}
.btn-start:disabled{background:var(--gray-200);color:var(--gray-400);cursor:not-allowed}
.btn-stop{background:#fee2e2;color:var(--red);display:none}
.btn-stop:hover{background:#fecaca}

/* ── Stats ── */
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px}
.stat{background:#fff;border:1px solid var(--gray-200);border-radius:var(--radius);padding:12px;text-align:center}
.stat-val{font-size:24px;font-weight:700}
.stat-lbl{font-size:11px;color:var(--gray-400);margin-top:2px}
.s-total .stat-val{color:var(--gray-800)}
.s-found .stat-val{color:var(--green)}
.s-notfound .stat-val{color:var(--red)}
.s-remain .stat-val{color:var(--blue)}

/* ── Progress ── */
.progress-wrap{height:6px;background:var(--gray-100);border-radius:3px;overflow:hidden;margin-bottom:14px;display:none}
.progress-fill{height:100%;background:var(--blue);border-radius:3px;transition:width .3s}

/* ── Filter tabs ── */
.tabs{display:flex;gap:6px;margin-bottom:10px;flex-wrap:wrap}
.tab{padding:5px 14px;border-radius:100px;border:1px solid var(--gray-200);font-size:12px;cursor:pointer;background:#fff;color:var(--gray-600);font-weight:500}
.tab.active{background:var(--blue);color:#fff;border-color:var(--blue)}

/* ── Table ── */
.table-wrap{max-height:420px;overflow-y:auto;border:1px solid var(--gray-200);border-radius:var(--radius-lg)}
table{width:100%;border-collapse:collapse;font-size:13px;table-layout:fixed}
th{text-align:left;padding:9px 12px;font-size:11px;font-weight:600;color:var(--gray-400);text-transform:uppercase;letter-spacing:.04em;border-bottom:1px solid var(--gray-200);background:#fff;position:sticky;top:0}
td{padding:9px 12px;border-bottom:1px solid #f5f5f5;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
tr:last-child td{border-bottom:none}
tr:hover td{background:var(--gray-50)}
.mono{font-family:monospace;font-size:12px}

/* ── Badges ── */
.badge{display:inline-flex;align-items:center;gap:3px;padding:2px 8px;border-radius:100px;font-size:11px;font-weight:600}
.b-found{background:var(--green-bg);color:var(--green);border:1px solid var(--green-bd)}
.b-nf{background:var(--red-bg);color:var(--red);border:1px solid var(--red-bd)}
.b-err{background:var(--yellow-bg);color:var(--yellow);border:1px solid var(--yellow-bd)}
.b-pend{background:var(--gray-100);color:var(--gray-400);border:1px solid var(--gray-200)}

/* ── Bottom toolbar ── */
.toolbar{display:flex;justify-content:space-between;align-items:center;margin-top:12px}
.toolbar-btns{display:flex;gap:8px}
.icon-btn{padding:7px 14px;border:1px solid var(--gray-200);border-radius:var(--radius);font-size:13px;font-weight:500;cursor:pointer;background:#fff;color:var(--gray-600);display:inline-flex;align-items:center;gap:6px}
.icon-btn:hover{background:var(--gray-50)}
.status-msg{font-size:12px;color:var(--gray-400)}

.empty{text-align:center;padding:48px 16px;color:var(--gray-400);font-size:13px}
.empty svg{margin-bottom:10px;opacity:.4}
</style>
</head>
<body>

<div class="topbar">
  <div class="topbar-brand">
    <svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 14l2 2 4-4"/><rect x="3" y="4" width="18" height="16" rx="2"/><path d="M7 4V2M17 4V2"/></svg>
    eBarimt VAT Checker
    <span class="topbar-sub">APU Trading</span>
  </div>
  <div style="font-size:12px;color:var(--gray-400)" id="clock"></div>
</div>

<div class="page">
  <div class="grid">

    <!-- Left panel -->
    <div>
      <div class="card">
        <div class="card-title">Input</div>

        <span class="field-label">Paste registration numbers</span>
        <textarea id="regInput" placeholder="5113377&#10;1234567&#10;9876543&#10;..."></textarea>

        <div class="or-divider">or upload a file</div>

        <label class="file-zone" id="dropZone">
          <input type="file" id="fileInput" accept=".txt,.csv">
          <div class="file-zone-text">
            <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="margin-bottom:4px;display:block;margin-left:auto;margin-right:auto"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
            <b>Choose file</b> or drag &amp; drop<br>.txt or .csv — one number per line
          </div>
          <div id="fileNameTag"></div>
        </label>
        <div id="countTag"></div>

        <div class="settings-row">
          <label>Parallel requests
            <select id="concurrency">
              <option value="1">1</option>
              <option value="3">3</option>
              <option value="5" selected>5</option>
              <option value="10">10</option>
            </select>
          </label>
          <label>Delay between calls
            <select id="delay">
              <option value="0">None</option>
              <option value="100">100 ms</option>
              <option value="200" selected>200 ms</option>
              <option value="500">500 ms</option>
            </select>
          </label>
        </div>

        <button class="btn btn-start" id="startBtn" onclick="startJob()">
          <svg width="16" height="16" viewBox="0 0 24 24" fill="currentColor"><polygon points="5 3 19 12 5 21 5 3"/></svg>
          Start verification
        </button>
        <button class="btn btn-stop" id="stopBtn" onclick="stopJob()">
          <svg width="16" height="16" viewBox="0 0 24 24" fill="currentColor"><rect x="3" y="3" width="18" height="18" rx="2"/></svg>
          Stop
        </button>
      </div>
    </div>

    <!-- Right panel -->
    <div>
      <div class="stats">
        <div class="stat s-total"><div class="stat-val" id="sTot">0</div><div class="stat-lbl">Total</div></div>
        <div class="stat s-found"><div class="stat-val" id="sOk">0</div><div class="stat-lbl">Found</div></div>
        <div class="stat s-notfound"><div class="stat-val" id="sErr">0</div><div class="stat-lbl">Not found</div></div>
        <div class="stat s-remain"><div class="stat-val" id="sPend">0</div><div class="stat-lbl">Remaining</div></div>
      </div>

      <div class="progress-wrap" id="progressWrap">
        <div class="progress-fill" id="progressFill" style="width:0%"></div>
      </div>

      <div class="tabs">
        <button class="tab active" onclick="setFilter('all',this)">All</button>
        <button class="tab" onclick="setFilter('Found',this)">Found</button>
        <button class="tab" onclick="setFilter('Not found',this)">Not found</button>
        <button class="tab" onclick="setFilter('Error',this)">Errors</button>
      </div>

      <div class="table-wrap">
        <table>
          <colgroup><col style="width:36px"><col style="width:120px"><col style="width:auto"><col style="width:90px"></colgroup>
          <thead><tr><th>#</th><th>Reg No</th><th>TIN</th><th>Status</th></tr></thead>
          <tbody id="tbody">
            <tr><td colspan="4"><div class="empty">
              <svg width="32" height="32" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>
              Paste numbers and click Start
            </div></td></tr>
          </tbody>
        </table>
      </div>

      <div class="toolbar">
        <span class="status-msg" id="statusMsg"></span>
        <div class="toolbar-btns">
          <button class="icon-btn" onclick="exportExcel()">
            <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
            Export Excel
          </button>
          <button class="icon-btn" onclick="clearAll()">
            <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polyline points="3 6 5 6 21 6"/><path d="M19 6l-1 14a2 2 0 01-2 2H8a2 2 0 01-2-2L5 6"/><path d="M10 11v6M14 11v6"/></svg>
            Clear
          </button>
        </div>
      </div>
    </div>

  </div>
</div>

<script>
let allResults = [], currentFilter = 'all', polling = null;

// Clock
setInterval(() => {
  const now = new Date();
  document.getElementById('clock').textContent =
    now.toLocaleDateString('mn-MN') + '  ' + now.toLocaleTimeString('mn-MN');
}, 1000);

// File input
document.getElementById('fileInput').addEventListener('change', e => {
  const f = e.target.files[0]; if(!f) return;
  document.getElementById('fileNameTag').textContent = f.name;
  const r = new FileReader();
  r.onload = ev => { document.getElementById('regInput').value = ev.target.result; updateCount(); };
  r.readAsText(f);
});
document.getElementById('regInput').addEventListener('input', updateCount);

function parseNums() {
  return [...new Set(
    document.getElementById('regInput').value
      .split(/[\n,;\s]+/).map(s=>s.trim()).filter(Boolean)
  )];
}
function updateCount() {
  const n = parseNums().length;
  document.getElementById('countTag').textContent = n ? n + ' numbers detected' : '';
}

async function startJob() {
  const nums = parseNums();
  if(!nums.length) { alert('Enter at least one registration number.'); return; }
  const conc  = document.getElementById('concurrency').value;
  const delay = document.getElementById('delay').value;

  document.getElementById('startBtn').style.display = 'none';
  document.getElementById('stopBtn').style.display  = '';
  document.getElementById('progressWrap').style.display = 'block';
  document.getElementById('statusMsg').textContent = 'Running...';
  allResults = [];
  renderTable();

  const res = await fetch('/start', {
    method: 'POST',
    headers: {'Content-Type':'application/json'},
    body: JSON.stringify({ numbers: nums, concurrency: parseInt(conc), delay: parseInt(delay) })
  });

  if(!res.ok) { alert('Failed to start job'); return; }
  polling = setInterval(pollStatus, 600);
}

async function stopJob() {
  await fetch('/stop', { method: 'POST' });
}

async function pollStatus() {
  try {
    const res = await fetch('/status');
    const data = await res.json();
    allResults = data.results;
    updateStats(data);
    renderTable();
    if(!data.running) {
      clearInterval(polling);
      document.getElementById('startBtn').style.display = '';
      document.getElementById('stopBtn').style.display  = 'none';
      const done = data.results.filter(r=>r.status!=='Pending').length;
      document.getElementById('statusMsg').textContent =
        'Done — ' + data.results.filter(r=>r.status==='Found').length + ' found out of ' + data.total;
    }
  } catch(e) {}
}

function updateStats(data) {
  const tot   = data.total || 0;
  const ok    = (data.results||[]).filter(r=>r.status==='Found').length;
  const err   = (data.results||[]).filter(r=>r.status==='Not found'||r.status==='Error').length;
  const pend  = (data.results||[]).filter(r=>r.status==='Pending').length;
  document.getElementById('sTot').textContent  = tot;
  document.getElementById('sOk').textContent   = ok;
  document.getElementById('sErr').textContent  = err;
  document.getElementById('sPend').textContent = pend;
  const done = tot - pend;
  document.getElementById('progressFill').style.width = tot ? Math.round(done/tot*100)+'%' : '0%';
}

function setFilter(f, el) {
  currentFilter = f;
  document.querySelectorAll('.tab').forEach(b=>b.classList.remove('active'));
  el.classList.add('active');
  renderTable();
}

function badge(s) {
  if(s==='Found')     return '<span class="badge b-found">✓ Found</span>';
  if(s==='Not found') return '<span class="badge b-nf">✗ Not found</span>';
  if(s==='Error')     return '<span class="badge b-err">⚠ Error</span>';
  return '<span class="badge b-pend">Pending</span>';
}

function renderTable() {
  const rows = currentFilter === 'all' ? allResults
    : allResults.filter(r => r.status === currentFilter);
  if(!rows.length) {
    document.getElementById('tbody').innerHTML =
      '<tr><td colspan="4"><div class="empty">No results yet</div></td></tr>';
    return;
  }
  document.getElementById('tbody').innerHTML = rows.map((r,i) => `
    <tr>
      <td style="color:#9ca3af">${i+1}</td>
      <td class="mono">${r.reg}</td>
      <td class="mono">${r.tin||'—'}</td>
      <td>${badge(r.status)}</td>
    </tr>`).join('');
}

async function exportExcel() {
  if(!allResults.length) { alert('No results to export.'); return; }
  window.location.href = '/export';
}

function clearAll() {
  allResults = [];
  document.getElementById('regInput').value = '';
  document.getElementById('countTag').textContent = '';
  document.getElementById('fileNameTag').textContent = '';
  document.getElementById('progressFill').style.width = '0%';
  document.getElementById('progressWrap').style.display = 'none';
  document.getElementById('statusMsg').textContent = '';
  ['sTot','sOk','sErr','sPend'].forEach(id=>document.getElementById(id).textContent='0');
  document.getElementById('tbody').innerHTML =
    '<tr><td colspan="4"><div class="empty"><svg width="32" height="32" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>Paste numbers and click Start</div></td></tr>';
  fetch('/clear', {method:'POST'});
}
</script>
</body>
</html>
"""


def fetch_reg(reg_no, delay_ms):
    """Call the eBarimt API for one reg number."""
    url = API_URL.format(quote(str(reg_no), safe=""))
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=10) as resp:
            body = resp.read().decode("utf-8")
        data = json.loads(body)
        time.sleep(delay_ms / 1000)
        if data.get("status") == 200 and data.get("data") not in (None, ""):
            return str(data["data"]), "Found"
        else:
            return "", "Not found"
    except Exception as e:
        return "", "Error"


def run_job(numbers, concurrency, delay_ms):
    """Run the batch job in background threads."""
    with job_lock:
        job["running"] = True
        job["total"] = len(numbers)
        job["done"] = 0
        job["results"] = [{"reg": n, "tin": "", "status": "Pending"} for n in numbers]

    idx_lock = threading.Lock()
    idx = [0]

    def worker():
        while True:
            with idx_lock:
                if idx[0] >= len(numbers):
                    break
                i = idx[0]
                idx[0] += 1
            with job_lock:
                if not job["running"]:
                    break
            tin, status = fetch_reg(numbers[i], delay_ms)
            with job_lock:
                job["results"][i] = {"reg": numbers[i], "tin": tin, "status": status}
                job["done"] += 1

    threads = [threading.Thread(target=worker, daemon=True) for _ in range(concurrency)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    with job_lock:
        job["running"] = False


def build_excel(results):
    """Build Excel file in memory and return bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VAT Results"

    hdr_fill = PatternFill("solid", fgColor="1a56db")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(["#", "Reg No", "TIN", "Status"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    green  = PatternFill("solid", fgColor="F0FDF4")
    red    = PatternFill("solid", fgColor="FEF2F2")
    yellow = PatternFill("solid", fgColor="FEFCE8")

    for i, r in enumerate(results, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r["reg"])
        ws.cell(row=row, column=3, value=r["tin"])
        ws.cell(row=row, column=4, value=r["status"])
        fill = green if r["status"] == "Found" else red if r["status"] == "Not found" else yellow
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = fill

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total   = len(results)
    found   = sum(1 for r in results if r["status"] == "Found")
    not_found = total - found
    ws2["A1"] = "VAT Verification Summary"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A3"] = "Run date";    ws2["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2["A4"] = "Total";       ws2["B4"] = total
    ws2["A5"] = "Found";       ws2["B5"] = found
    ws2["A5"].font = Font(color="15803d"); ws2["B5"].font = Font(color="15803d")
    ws2["A6"] = "Not found";   ws2["B6"] = not_found
    ws2["A6"].font = Font(color="b91c1c"); ws2["B6"].font = Font(color="b91c1c")
    ws2["A7"] = "Match rate";  ws2["B7"] = f"{found/total*100:.1f}%" if total else "0%"
    ws2["A7"].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass  # silence default access log

    def send_json(self, data, status=200):
        body = json.dumps(data).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        path = urlparse(self.path).path

        if path == "/" or path == "/index.html":
            body = HTML.encode()
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", len(body))
            self.end_headers()
            self.wfile.write(body)

        elif path == "/status":
            with job_lock:
                self.send_json({
                    "running": job["running"],
                    "total":   job["total"],
                    "done":    job["done"],
                    "results": job["results"],
                })

        elif path == "/export":
            with job_lock:
                results = list(job["results"])
            if not results:
                self.send_json({"error": "No results"}, 400)
                return
            xlsx = build_excel(results)
            fname = f"vat_results_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
            self.send_header("Content-Length", len(xlsx))
            self.end_headers()
            self.wfile.write(xlsx)

        else:
            self.send_response(404); self.end_headers()

    def do_POST(self):
        path = urlparse(self.path).path
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length)

        if path == "/start":
            with job_lock:
                if job["running"]:
                    self.send_json({"error": "Already running"}, 409)
                    return
            payload = json.loads(body)
            numbers     = payload.get("numbers", [])
            concurrency = max(1, min(10, int(payload.get("concurrency", 5))))
            delay_ms    = max(0, int(payload.get("delay", 200)))
            threading.Thread(target=run_job, args=(numbers, concurrency, delay_ms), daemon=True).start()
            self.send_json({"ok": True})

        elif path == "/stop":
            with job_lock:
                job["running"] = False
            self.send_json({"ok": True})

        elif path == "/clear":
            with job_lock:
                job["results"] = []
                job["total"] = 0
                job["done"] = 0
                job["running"] = False
            self.send_json({"ok": True})

        else:
            self.send_response(404); self.end_headers()


def main():
    server = HTTPServer(("0.0.0.0", PORT), Handler)
    url = f"http://localhost:{PORT}"

    print(f"\n  eBarimt VAT Checker running at {url}")
    print("")
    print("  Press Ctrl+C to stop\n")


    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Stopped.")


if __name__ == "__main__":
    main()
